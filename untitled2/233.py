# coding:utf-8
from openpyxl import load_workbook
from openpyxl import Workbook
import requests
from docopt import docopt
from bs4 import BeautifulSoup
"""Phoneaddress
Usage:
    searchphone <telephone>
    searchfile <file_name> <file_sheet>
"""

arguments =docopt(__doc__, version='Phoneaddress 0.0.1')
def load_file():
    wb = load_workbook(filename='33.xlsx')
    ws = wb.get_sheet_by_name('Sheet1')
    rows = ws.rows
    columns = ws.columns
    print(len(rows))
    a=[]
    for i in range(1,(len(rows)+1)):
        t= ws.cell(row = i, column = 1).value
        a.append(t)
    return a

def save_file(a,b):
    print(len(a))
    wb = Workbook()
    new_ws = wb.create_sheet(title='2333')
    for rows in range(1,(len(a)+1)):
        new_ws.cell(row=rows, column=1).value = a[rows-1]
    for rows in range(1,(len(b)+1)):
        new_ws.cell(row=rows, column=2).value = b[rows-1]
    wb.save(filename='RESULT2.xlsx')

def find(telephone):
    url = 'http://www.ip138.com:8080/search.asp?action=mobile&mobile='
    print(telephone)
    tele=str(telephone)
    r=requests.get(url+tele)
    r.encoding='gbk'
    context=r.text
    soup=BeautifulSoup(context, "lxml")
    try:
        data = soup.find_all('td')[6].text.split('\xa0')
        request = data[0] + ' ' + data[1]
    except IndexError:
        request='ERROR'
    return request

def main():
    tel=[]
    address=[]
    tel=load_file()
    for i in tel:
        address.append(find(i))
    save_file(tel,address)

if __name__ == '__main__':
    main()


