#coding:utf-8
from openpyxl import load_workbook
class find_aderess():
    def data_load(self,a):
        self.a=a
        wb = load_workbook(filename=r'33.xlsx')
        sheets = wb.get_sheet_names()  # 获取所有表格(worksheet)的名字
        sheet1 = sheets[0]  # 第一个表格的名称
        ws = wb.get_sheet_by_name('Sheet1')  # 获取特定的 worksheet
        a=ws.cell('A1').value
        print(self.a)
    def data_save(self):
        pass
    def choose(self):
        pass

